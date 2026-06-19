# backend/app/services/excel_processor.py

import os
import re
import logging
import openpyxl
from datetime import datetime

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Serviço para processar arquivos Excel (.xlsx) de rondas."""

    @staticmethod
    def parse_excel_file(filepath: str) -> dict:
        """
        Carrega o arquivo Excel e extrai as informações de cabeçalho e rondas por condomínio.
        """
        if not os.path.exists(filepath):
            logger.error(f"Arquivo não encontrado: {filepath}")
            return {"success": False, "message": "Arquivo não encontrado."}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "Rondas" not in wb.sheetnames:
                logger.warning(f"Planilha 'Rondas' não encontrada no arquivo {filepath}.")
                return {"success": False, "message": "Planilha 'Rondas' não encontrada no arquivo Excel."}

            sheet = wb["Rondas"]

            # 1. Parse de informações do cabeçalho
            supervisor = None
            turno = None
            data_plantao = None

            # Varre as primeiras 5 linhas em busca das informações básicas do plantão
            for i in range(1, 6):
                for col in range(1, 6):
                    val = sheet.cell(row=i, column=col).value
                    if val and isinstance(val, str):
                        if "Supervisor:" in val:
                            sup_match = re.search(r"Supervisor:\s*([^|]+)", val)
                            if sup_match:
                                supervisor = sup_match.group(1).strip()
                        if "Turno:" in val:
                            turno_match = re.search(r"Turno:\s*([^|]+)", val)
                            if turno_match:
                                turno = turno_match.group(1).strip()
                        if "Data:" in val:
                            data_match = re.search(r"Data:\s*(\d{2}/\d{2}/\d{4})", val)
                            if data_match:
                                data_plantao = data_match.group(1).strip()

            # Normalização de escala_plantao baseada no turno
            escala_plantao = "18h às 06h"  # Default
            if turno:
                turno_lower = turno.lower()
                if "diurno" in turno_lower or "06h" in turno_lower or ("06:00" in turno_lower and "18:00" not in turno_lower):
                    escala_plantao = "06h às 18h"
                elif "noturno" in turno_lower or "18h" in turno_lower or "18:00" in turno_lower:
                    escala_plantao = "18h às 06h"

            # Conversão de data para formato ISO (YYYY-MM-DD)
            data_iso = None
            if data_plantao:
                try:
                    dt = datetime.strptime(data_plantao, "%d/%m/%Y")
                    data_iso = dt.strftime("%Y-%m-%d")
                except ValueError:
                    pass

            # 2. Parse de condomínios e suas rondas
            condominios_data = {}
            current_condo = None

            for row_idx in range(1, sheet.max_row + 1):
                cell_val = sheet.cell(row=row_idx, column=1).value
                if cell_val and isinstance(cell_val, str) and cell_val.strip().lower().startswith("residencial:"):
                    # Identifica novo residencial
                    condo_name = cell_val.replace("Residencial:", "").replace("residencial:", "").strip()
                    current_condo = condo_name
                    condominios_data[current_condo] = []
                    continue

                if current_condo:
                    col1 = sheet.cell(row=row_idx, column=1).value
                    if col1 and isinstance(col1, str) and col1.strip().lower().startswith("ronda"):
                        # Captura as informações da rodada
                        round_name = col1.strip()
                        inicio = sheet.cell(row=row_idx, column=2).value
                        termino = sheet.cell(row=row_idx, column=3).value
                        duracao = sheet.cell(row=row_idx, column=4).value
                        vtr_agente = sheet.cell(row=row_idx, column=5).value

                        condominios_data[current_condo].append({
                            "round_name": round_name,
                            "inicio": str(inicio).strip() if inicio else None,
                            "termino": str(termino).strip() if termino else None,
                            "duracao": str(duracao).strip() if duracao else None,
                            "vtr_agente": str(vtr_agente).strip() if vtr_agente else None
                        })
                    elif col1 and isinstance(col1, str) and "total" in col1.strip().lower():
                        # Fim das rondas para esse residencial
                        current_condo = None

            return {
                "success": True,
                "supervisor": supervisor,
                "turno": turno,
                "escala_plantao": escala_plantao,
                "data_plantao": data_plantao,
                "data_iso": data_iso,
                "condominios": condominios_data
            }

        except Exception as e:
            logger.error(f"Erro ao analisar o arquivo Excel {filepath}: {e}", exc_info=True)
            return {"success": False, "message": f"Erro ao processar arquivo Excel: {str(e)}"}

    @staticmethod
    def generate_simulated_whatsapp_log(parsed_data: dict, condominio_name: str) -> str:
        """
        Gera o log bruto simulado do WhatsApp para um condomínio específico.
        """
        if not parsed_data or not parsed_data.get("success") or not condominio_name:
            return ""

        target_condo = None
        for name in parsed_data.get("condominios", {}):
            if name.strip().lower() == condominio_name.strip().lower():
                target_condo = name
                break

        if not target_condo:
            logger.warning(f"Condomínio '{condominio_name}' não encontrado nos dados extraídos do Excel.")
            return ""

        lines = []
        date_str = parsed_data.get("data_plantao") or datetime.now().strftime("%d/%m/%Y")

        for r in parsed_data["condominios"][target_condo]:
            # Extrai os dígitos identificadores do veículo (ex: MT-03 -> VTR 03)
            vtr_num = "01"
            if r.get("vtr_agente"):
                vtr_match = re.search(r'\d+', r["vtr_agente"])
                if vtr_match:
                    vtr_num = vtr_match.group(0).zfill(2)

            inicio = r.get("inicio")
            termino = r.get("termino")

            if inicio and inicio != '--':
                lines.append(f"[{inicio}, {date_str}] VTR {vtr_num}: {inicio} inicio de ronda")
            if termino and termino != '--':
                lines.append(f"[{termino}, {date_str}] VTR {vtr_num}: {termino} termino de ronda")

        return "\n".join(lines)

    @staticmethod
    def parse_excel_file_paradas(filepath: str) -> dict:
        """
        Carrega o arquivo Excel e extrai as informações de cabeçalho e paradas por condomínio da aba 'Paradas'.
        """
        if not os.path.exists(filepath):
            logger.error(f"Arquivo não encontrado: {filepath}")
            return {"success": False, "message": "Arquivo não encontrado."}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "Paradas" not in wb.sheetnames:
                logger.warning(f"Planilha 'Paradas' não encontrada no arquivo {filepath}.")
                return {"success": False, "message": "Planilha 'Paradas' não encontrada no arquivo Excel."}

            sheet = wb["Paradas"]

            # 1. Parse de informações do cabeçalho
            supervisor = None
            turno = None
            data_plantao = None

            for i in range(1, 6):
                for col in range(1, 6):
                    val = sheet.cell(row=i, column=col).value
                    if val and isinstance(val, str):
                        if "Supervisor:" in val:
                            sup_match = re.search(r"Supervisor:\s*([^|]+)", val)
                            if sup_match:
                                supervisor = sup_match.group(1).strip()
                        if "Turno:" in val:
                            turno_match = re.search(r"Turno:\s*([^|]+)", val)
                            if turno_match:
                                turno = turno_match.group(1).strip()
                        if "Data:" in val:
                            data_match = re.search(r"Data:\s*(\d{2}/\d{2}/\d{4})", val)
                            if data_match:
                                data_plantao = data_match.group(1).strip()

            escala_plantao = "18h às 06h"
            if turno:
                turno_lower = turno.lower()
                if "diurno" in turno_lower or "06h" in turno_lower or ("06:00" in turno_lower and "18:00" not in turno_lower):
                    escala_plantao = "06h às 18h"
                elif "noturno" in turno_lower or "18h" in turno_lower or "18:00" in turno_lower:
                    escala_plantao = "18h às 06h"

            data_iso = None
            if data_plantao:
                try:
                    dt = datetime.strptime(data_plantao, "%d/%m/%Y")
                    data_iso = dt.strftime("%Y-%m-%d")
                except ValueError:
                    pass

            # 2. Parse de condomínios e suas paradas
            condominios_data = {}
            current_condo = None

            for row_idx in range(1, sheet.max_row + 1):
                cell_val = sheet.cell(row=row_idx, column=1).value
                if cell_val and isinstance(cell_val, str) and cell_val.strip().lower().startswith("residencial:"):
                    condo_name = cell_val.replace("Residencial:", "").replace("residencial:", "").strip()
                    current_condo = condo_name
                    condominios_data[current_condo] = []
                    continue

                if current_condo:
                    col1 = sheet.cell(row=row_idx, column=1).value
                    if col1 and isinstance(col1, str) and col1.strip().lower().startswith("parada"):
                        round_name = col1.strip()
                        inicio = sheet.cell(row=row_idx, column=2).value
                        termino = sheet.cell(row=row_idx, column=3).value
                        duracao = sheet.cell(row=row_idx, column=4).value
                        vtr_agente = sheet.cell(row=row_idx, column=5).value

                        condominios_data[current_condo].append({
                            "round_name": round_name,
                            "inicio": str(inicio).strip() if inicio else None,
                            "termino": str(termino).strip() if termino else None,
                            "duracao": str(duracao).strip() if duracao else None,
                            "vtr_agente": str(vtr_agente).strip() if vtr_agente else None
                        })
                    elif col1 and isinstance(col1, str) and "total" in col1.strip().lower():
                        current_condo = None

            return {
                "success": True,
                "supervisor": supervisor,
                "turno": turno,
                "escala_plantao": escala_plantao,
                "data_plantao": data_plantao,
                "data_iso": data_iso,
                "condominios": condominios_data
            }

        except Exception as e:
            logger.error(f"Erro ao analisar o arquivo Excel {filepath}: {e}", exc_info=True)
            return {"success": False, "message": f"Erro ao processar arquivo Excel: {str(e)}"}

    @staticmethod
    def generate_simulated_whatsapp_log_parada(parsed_data: dict, condominio_name: str) -> str:
        """
        Gera o log bruto simulado do WhatsApp para um condomínio específico (Paradas).
        """
        if not parsed_data or not parsed_data.get("success") or not condominio_name:
            return ""

        target_condo = None
        for name in parsed_data.get("condominios", {}):
            if name.strip().lower() == condominio_name.strip().lower():
                target_condo = name
                break

        if not target_condo:
            logger.warning(f"Condomínio '{condominio_name}' não encontrado nos dados extraídos do Excel.")
            return ""

        lines = []
        date_str = parsed_data.get("data_plantao") or datetime.now().strftime("%d/%m/%Y")

        for r in parsed_data["condominios"][target_condo]:
            vtr_num = "01"
            if r.get("vtr_agente"):
                vtr_match = re.search(r'\d+', r["vtr_agente"])
                if vtr_match:
                    vtr_num = vtr_match.group(0).zfill(2)

            inicio = r.get("inicio")
            termino = r.get("termino")

            if inicio and inicio != '--':
                lines.append(f"[{inicio}, {date_str}] VTR {vtr_num}: {inicio} inicio de parada")
            if termino and termino != '--':
                lines.append(f"[{termino}, {date_str}] VTR {vtr_num}: {termino} termino de parada")

        return "\n".join(lines)

